#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PDF zu CSV - Automatisierung der IHK-Fragengenerierung
Konvertiert jede Seite eines PDFs in Single-Choice-Fragen via OpenAI (Vision).

Verwendung:
    # Standard (Doppelseiten: 1 PDF = 2 Buchseiten)
    python pdf_to_csv.py --pdf buch.pdf --prompt megaprompt.docx --start 8 --end 182

    # Einzelseiten (1 PDF = 1 Buchseite)
    python pdf_to_csv.py --pdf buch.pdf --prompt megaprompt.docx --start 8 --end 182 --single-page

    # Mit spezifischer Buchseitennummer (PDF-Seite 8 = Buchseite 15)
    python pdf_to_csv.py --pdf buch.pdf --prompt megaprompt.docx --start 8 --end 182 --book-start 15

    # Vollständiges Beispiel
    python pdf_to_csv.py --pdf Buch.pdf --prompt Megaprompt.docx --model gpt-5.1 --start 8 --end 182 --parallel 3 --single-page --book-start 15

Optionen:
    --pdf           Pfad zur PDF-Datei
    --prompt        Pfad zur Megaprompt DOCX-Datei
    --start         Erste PDF-Seite (Standard: 1)
    --end           Letzte PDF-Seite (Standard: letzte)
    --model         gpt-4.1, gpt-5.1, gpt-5.2 (Standard: gpt-4.1)
    --parallel      Anzahl paralleler Anfragen (Standard: 3)
    --single-page   1 PDF-Seite = 1 Buchseite (Standard: Doppelseiten)
    --book-start    Buchseitennummer für erste PDF-Seite
    --test          Nur 2 Seiten verarbeiten
"""

import os
import sys
import argparse
import base64
import re
import time
import json
import random
import signal
from pathlib import Path
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock, Event

import fitz  # PyMuPDF
from docx import Document
from openai import OpenAI
from dotenv import load_dotenv

# Umgebungsvariablen laden
load_dotenv()

# Konfiguration
MAX_VERSUCHE = 5
BASIS_WARTEZEIT = 2
MAX_WARTEZEIT = 60
DPI = 150
JPEG_QUALITAET = 85  # JPEG-Qualität (schnellere Übertragung als PNG)

# CSV-Spaltenüberschriften (gemäß Megaprompt)
CSV_HEADER = "Frage;A;B;C;D;Richtig;Richtig_Text;Thema;Quelle;Status;Kommentar;Vollansicht"
EXPECTED_COLS = 12

# Erkannte Fehlertypen
FEHLER_TYPEN = {
    "CONTENT_POLICY": "API hat den Inhalt abgelehnt (Inhaltsrichtlinie)",
    "KEIN_TEXT": "Die Seite enthält nicht genug verwertbaren Text",
    "NUR_BILD": "Die Seite enthält nur Bilder/Diagramme",
    "UNGUELTIG_CSV": "Die Antwort ist nicht im gültigen CSV-Format",
    "API_FEHLER": "Technischer API-Fehler",
    "VERBINDUNG": "Verbindungsfehler - Netzwerk oder Server nicht erreichbar",
    "RATE_LIMIT": "API-Anfragelimit erreicht - automatische Wartezeit",
    "ZEITÜBERSCHREITUNG": "Zeitlimit überschritten",
    "UNBEKANNT": "Unbekannter Fehler"
}

# Thread-sicheres Schreiben
schreib_lock = Lock()

# Globale Variable für den geladenen Megaprompt
MEGAPROMPT_INHALT = None

# Flag für graceful shutdown bei Ctrl+C
stop_event = Event()
wurde_unterbrochen = False

# Progress-Tracking (Resume-System)
PROGRESS_PFAD = None
ERLEDIGTE_SEITEN = set()

# Metriken-Sammlung (Étape 2)
METRIKEN_LISTE = []


def lade_progress(progress_pfad: str) -> set:
    """Lädt bereits verarbeitete Seiten aus dem Progress-File."""
    if os.path.exists(progress_pfad):
        try:
            with open(progress_pfad, "r", encoding="utf-8") as f:
                data = json.load(f)
                return set(data.get("erledigte_seiten", []))
        except Exception:
            return set()
    return set()


def speichere_progress(progress_pfad: str, seite: int):
    """Fügt eine erledigte Seite zum Progress-File hinzu (thread-sicher)."""
    with schreib_lock:
        ERLEDIGTE_SEITEN.add(seite)
        with open(progress_pfad, "w", encoding="utf-8") as f:
            json.dump({
                "erledigte_seiten": sorted(list(ERLEDIGTE_SEITEN)),
                "letzte_aktualisierung": datetime.now().isoformat()
            }, f, ensure_ascii=False, indent=2)


def loesche_progress(progress_pfad: str):
    """Löscht das Progress-File nach erfolgreichem Abschluss."""
    if os.path.exists(progress_pfad):
        os.remove(progress_pfad)
        print("   ✓ Progress-Datei gelöscht (Verarbeitung vollständig)")


def signal_handler(sig, frame):
    """Handler für Ctrl+C - setzt Stop-Flag statt sofort zu beenden."""
    global wurde_unterbrochen
    wurde_unterbrochen = True
    stop_event.set()
    print("\n\n⚠️  UNTERBRECHUNG ERKANNT - Speichere bisherige Daten...")
    print("   (Bitte warten, laufende Anfragen werden abgeschlossen)")


def schreibe_ergebnis_sofort(ausgabe_pfad: str, ergebnis: dict, statistik: dict):
    """Schreibt ein Ergebnis sofort in die CSV-Datei und aktualisiert den Progress (thread-sicher)."""
    with schreib_lock:
        if not ergebnis["fehler"] and ergebnis["csv_inhalt"]:
            with open(ausgabe_pfad, "a", encoding="utf-8-sig", newline="") as csv_datei:
                csv_datei.write(ergebnis["csv_inhalt"] + "\n")
            statistik["erfolg"] += 1
            statistik["fragen_generiert"] += ergebnis["anzahl_fragen"]
            # Progress speichern
            if PROGRESS_PFAD:
                ERLEDIGTE_SEITEN.add(ergebnis["pdf_seite"])
                with open(PROGRESS_PFAD, "w", encoding="utf-8") as f:
                    json.dump({
                        "erledigte_seiten": sorted(list(ERLEDIGTE_SEITEN)),
                        "letzte_aktualisierung": datetime.now().isoformat()
                    }, f, ensure_ascii=False, indent=2)


def lade_megaprompt(docx_pfad: str) -> str:
    """Lädt den Megaprompt aus einer DOCX-Datei."""
    if not os.path.exists(docx_pfad):
        raise FileNotFoundError(f"Megaprompt-Datei nicht gefunden: {docx_pfad}")

    doc = Document(docx_pfad)
    text = "\n".join([p.text for p in doc.paragraphs])
    return text


def analysiere_und_konvertiere_seite(pdf_doc, seiten_nr: int, dpi: int = DPI) -> tuple[dict, str]:
    """
    Kombinierte Funktion: Analysiert die Seite UND konvertiert sie zu Base64.
    Vermeidet doppeltes Öffnen des PDFs -> schneller.

    Rückgabe: (meta_dict, base64_bild)
    """
    page = pdf_doc[seiten_nr]

    # Analyse
    text = page.get_text("text") or ""
    images = page.get_images(full=True) or []
    text_norm = re.sub(r"\s+", "", text)
    text_len = len(text_norm)

    meta = {
        "text_len": text_len,
        "image_count": len(images),
        "ist_bildlastig": (text_len < 80 and len(images) > 0),
        "ist_textlastig": (text_len >= 200),
    }

    # Konvertierung zu JPEG (kleiner als PNG, schnellere Übertragung)
    mat = fitz.Matrix(dpi / 72, dpi / 72)
    pix = page.get_pixmap(matrix=mat)
    bild_bytes = pix.tobytes("jpeg")  # JPEG statt PNG
    base64_bild = base64.b64encode(bild_bytes).decode("utf-8")

    return meta, base64_bild


def get_megaprompt_mit_quelle(seiten_info: str) -> str:
    """Gibt den geladenen Megaprompt mit der Quellenangabe zurück."""
    global MEGAPROMPT_INHALT
    if not MEGAPROMPT_INHALT:
        raise ValueError("Megaprompt wurde nicht geladen!")

    return f"""{MEGAPROMPT_INHALT}

---
AKTUELLE QUELLE für diese Doppelseite: {seiten_info}

WICHTIG:
- Setze bei jeder Frage in der Spalte "Quelle" den Wert: {seiten_info}
- Setze bei jeder Frage in der Spalte "Status" den Wert: ok
- KEINE Semikolons (;) im Textinhalt der Fragen/Antworten!
- Wenn du ein Semikolon verwenden würdest: ersetze es durch ein Komma oder einen Gedankenstrich
- Verwende innerhalb von Feldern niemals Zeilenumbrüche

QUALITÄTS-CHECK (NICHT AUSGEBEN):
- exakt 12 Spalten je Zeile (11 Semikolons)
- Richtig ist A/B/C/D
- Richtig_Text entspricht exakt der gewählten Option (A/B/C/D)
- keine Einleitung, keine Überschriften, keine Erklärungen außerhalb CSV

Beginne SOFORT mit der ersten CSV-Zeile, KEINE Einleitung!
"""


def get_standard_prompt(seiten_info: str) -> str:
    """Fallback-Prompt falls Megaprompt nicht funktioniert."""
    return f"""Erstelle 12-15 Single-Choice-Prüfungsfragen für IHK-Einzelhandel.

Regeln:
- 4 Antworten pro Frage (A-D), eine richtig
- Richtige Antwort zufällig verteilen
- KEINE Semikolons im Text (ersetzte sie durch Komma oder Gedankenstrich)
- Keine Zeilenumbrüche innerhalb eines Feldes
- Praxisnah und verständlich
- Gendergerechte Du-Form
- Keine Einleitung, keine Erklärungen

Fallbeispiele mit: Frau Rabatta (Modehaus), Herr Andreh (Frischemarkt), Azubi Lisa, Azubi Mehmet

CSV-Format: Semikolon-getrennt, OHNE Kopfzeile, exakt 12 Spalten:
Frage;A;B;C;D;Richtig;Richtig_Text;Thema;Quelle;Status;Kommentar;Vollansicht

Quelle: {seiten_info}
Status: ok

QUALITÄT (nicht ausgeben):
- 12 Spalten je Zeile
- Richtig_Text == exakt Option

Nur CSV ausgeben, beginne sofort:"""


def get_bildlastig_prompt(seiten_info: str) -> str:
    """
    Besserer Prompt für diagramm-/tabellenlastige Seiten.
    """
    return f"""Du siehst eine Buchseite mit wenig Fließtext, aber ggf. Tabellen, Diagrammen, Schaubildern oder Prozessgrafiken.

Aufgabe:
- Interpretiere Tabellen/Diagramme (Achsen, Legenden, Wertebereiche, Kernaussagen).
- Erstelle 10-12 IHK-relevante Single-Choice-Fragen dazu.
- Vermeide Formulierungen wie „siehe Abbildung“ oder „wie in der Grafik“.

CSV-Regeln:
- exakt 12 Spalten, Semikolon als Trennzeichen
- In Feldtexten niemals Semikolon und niemals Zeilenumbrüche
- Spalten: Frage;A;B;C;D;Richtig;Richtig_Text;Thema;Quelle;Status;Kommentar;Vollansicht
- Quelle={seiten_info}
- Status=ok

QUALITÄT (nicht ausgeben):
- 12 Spalten je Zeile
- Richtig_Text == exakt Option
- Keine Einleitung

Beginne sofort mit der ersten CSV-Zeile:"""


def erkenne_fehlertyp(antwort: str, ausnahme: Exception = None) -> tuple[str, str]:
    """Erkennt den Fehlertyp und gibt (Code, explizite Nachricht) zurück."""
    if ausnahme:
        fehler_text = str(ausnahme).lower()
        exception_typ = type(ausnahme).__name__.lower()
        alle_texte = f"{fehler_text} {exception_typ}"

        if ausnahme.__cause__:
            alle_texte += f" {str(ausnahme.__cause__).lower()} {type(ausnahme.__cause__).__name__.lower()}"
        if ausnahme.__context__:
            alle_texte += f" {str(ausnahme.__context__).lower()} {type(ausnahme.__context__).__name__.lower()}"

        verbindungs_keywords = [
            "connection", "connect", "network", "socket", "refused",
            "reset", "broken", "pipe", "eof", "ssl", "handshake",
            "apiconnection", "connectionerror", "remotedisconnected",
            "newconnectionerror", "maxretryerror", "urlerror"
        ]
        if any(w in alle_texte for w in verbindungs_keywords):
            return "VERBINDUNG", f"{FEHLER_TYPEN['VERBINDUNG']} - Detail: {str(ausnahme)[:100]}"

        if any(w in alle_texte for w in ["rate", "limit", "429", "quota", "exceeded"]):
            return "RATE_LIMIT", FEHLER_TYPEN["RATE_LIMIT"]

        if any(w in alle_texte for w in ["timeout", "timed out", "deadline"]):
            return "ZEITÜBERSCHREITUNG", FEHLER_TYPEN["ZEITÜBERSCHREITUNG"]

        if any(w in alle_texte for w in ["api", "invalid", "401", "403", "500", "502", "503", "badrequest"]):
            return "API_FEHLER", FEHLER_TYPEN["API_FEHLER"]

    if antwort:
        antwort_klein = antwort.lower()

        if any(phrase in antwort_klein for phrase in [
            "i can't assist", "i cannot assist", "i'm sorry", "i am sorry",
            "can't help", "cannot help", "not able to", "unable to",
            "policy", "guidelines", "inappropriate"
        ]):
            return "CONTENT_POLICY", FEHLER_TYPEN["CONTENT_POLICY"]

    # Wenn Antwort vorhanden, aber kaum Semikolons -> sehr wahrscheinlich ungültig
    if antwort and ';' not in antwort:
        return "UNGUELTIG_CSV", FEHLER_TYPEN["UNGUELTIG_CSV"]

    return "UNBEKANNT", FEHLER_TYPEN["UNBEKANNT"]


def berechne_wartezeit(versuch: int, fehler_typ: str) -> float:
    """Exponentieller Backoff mit Jitter."""
    if fehler_typ == "RATE_LIMIT":
        basis = BASIS_WARTEZEIT * 4
    elif fehler_typ == "VERBINDUNG":
        basis = BASIS_WARTEZEIT * 2
    else:
        basis = BASIS_WARTEZEIT

    wartezeit = min(basis * (2 ** versuch), MAX_WARTEZEIT)
    jitter = wartezeit * 0.2 * (random.random() * 2 - 1)
    return max(1, wartezeit + jitter)


def hole_pdf_seitenanzahl(pdf_pfad: str) -> int:
    """Gibt die Anzahl der Seiten im PDF zurück."""
    doc = fitz.open(pdf_pfad)
    anzahl = len(doc)
    doc.close()
    return anzahl


def rufe_openai_vision(client: OpenAI, base64_bild: str, prompt: str, modell: str = "gpt-4.1") -> str:
    """Ruft die OpenAI Vision API mit dem Bild auf."""
    # GPT-5 Modelle verwenden max_completion_tokens statt max_tokens
    ist_gpt5 = modell.startswith("gpt-5")

    api_params = {
        "model": modell,
        "messages": [
            {
                "role": "system",
                "content": (
                    "Du bist ein Experte für IHK-Prüfungsfragen. "
                    "Antworte IMMER und NUR im CSV-Format. "
                    "Keine Einleitungen, keine Erklärungen."
                )
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_bild}",
                            "detail": "high"
                        }
                    }
                ]
            }
        ],
        "temperature": 0.7,
        "timeout": 120
    }

    # Richtigen Token-Parameter je nach Modell verwenden
    if ist_gpt5:
        api_params["max_completion_tokens"] = 4096
    else:
        api_params["max_tokens"] = 4096

    antwort = client.chat.completions.create(**api_params)
    return antwort.choices[0].message.content, antwort.usage


def bereinige_und_validiere_csv(antwort: str) -> tuple[str, int, list]:
    """
    Bereinigt die API-Antwort, validiert streng (exakt 12 Spalten)
    und gibt (csv, anzahl_fragen, fehler) zurück.
    """
    if not antwort:
        return "", 0, ["Leere Antwort"]

    # Markdown-Codeblöcke entfernen
    antwort = re.sub(r"```(?:csv)?\s*", "", antwort)
    antwort = re.sub(r"```\s*", "", antwort)

    zeilen = [z.strip() for z in antwort.splitlines() if z.strip()]
    gueltige_zeilen = []
    fehler = []

    for i, zeile in enumerate(zeilen, 1):
        if not zeile or zeile.startswith("#"):
            continue
        if ";" not in zeile:
            continue

        teile = zeile.split(";")

        if len(teile) != EXPECTED_COLS:
            fehler.append(f"Zeile {i}: {len(teile)} Spalten (erwartet {EXPECTED_COLS})")
            continue

        teile = [t.strip() for t in teile]

        # Frage darf nicht leer sein
        if not teile[0]:
            fehler.append(f"Zeile {i}: Leere Frage")
            continue

        # Antworten A-D müssen vorhanden sein
        for j in range(1, 5):
            if not teile[j]:
                fehler.append(f"Zeile {i}: Antwort {chr(64 + j)} ist leer")
                break
        else:
            richtig = teile[5].upper()
            if richtig not in ["A", "B", "C", "D"]:
                fehler.append(f"Zeile {i}: Ungültige Antwort '{teile[5]}' (A/B/C/D erwartet)")
                continue
            teile[5] = richtig

            # Richtig_Text muss exakt der gewählten Option entsprechen
            idx = {"A": 1, "B": 2, "C": 3, "D": 4}[richtig]
            if teile[6] != teile[idx]:
                fehler.append(f"Zeile {i}: Richtig_Text passt nicht zu Option {richtig}")
                continue

            # Semikolon ist als Trennzeichen ok, aber im Inhalt darf keins stecken.
            # Da wir splitten, wäre ein zusätzliches ; schon an Spaltenzahl gescheitert.

            gueltige_zeilen.append(";".join(teile))

    return "\n".join(gueltige_zeilen), len(gueltige_zeilen), fehler


def protokolliere_fehler(
    fehler_log_pfad: str,
    pdf_seite: int,
    buch_seiten_info: str,
    fehler_typ: str,
    fehler_nachricht: str,
    rohe_antwort: str = None
):
    """Protokolliert einen Fehler in der Log-Datei (thread-sicher)."""
    zeitstempel = datetime.now().isoformat()

    fehler_eintrag = {
        "zeitstempel": zeitstempel,
        "pdf_seite": pdf_seite,
        "buch_seiten": buch_seiten_info,
        "fehler_typ": fehler_typ,
        "fehler_nachricht": fehler_nachricht,
        "rohe_antwort_vorschau": rohe_antwort[:500] if rohe_antwort else None
    }

    with schreib_lock:
        # JSON-Log
        fehler = []
        if os.path.exists(fehler_log_pfad):
            try:
                with open(fehler_log_pfad, "r", encoding="utf-8") as f:
                    fehler = json.load(f)
            except Exception:
                fehler = []

        fehler.append(fehler_eintrag)

        with open(fehler_log_pfad, "w", encoding="utf-8") as f:
            json.dump(fehler, f, ensure_ascii=False, indent=2)

        # Text-Log
        txt_log_pfad = fehler_log_pfad.replace(".json", ".txt")
        with open(txt_log_pfad, "a", encoding="utf-8") as f:
            f.write(f"\n{'=' * 60}\n")
            f.write(f"[{zeitstempel}] PDF-Seite {pdf_seite} ({buch_seiten_info})\n")
            f.write(f"Typ: {fehler_typ}\n")
            f.write(f"Nachricht: {fehler_nachricht}\n")


def verarbeite_seite(
    client: OpenAI,
    pdf_doc,
    seiten_nr: int,
    seiten_info: str,
    modell: str = "gpt-4.1",
    fehler_log_pfad: str = None
) -> tuple[str, str, str, int, dict]:
    """
    Verarbeitet eine PDF-Seite mit Retry und Prompt-Strategien.
    Rückgabe: (csv_inhalt, fehler_nachricht, fehler_typ, anzahl_fragen, metriken)
    """
    letzte_antwort = None
    seite_startzeit = time.time()
    api_aufrufe = 0

    metriken = {
        "strategie": None,
        "api_aufrufe": 0,
        "tokens_eingabe": 0,
        "tokens_ausgabe": 0,
        "tokens_gesamt": 0,
        "antwortzeit_s": 0.0,
        "gesamtzeit_s": 0.0,
        "fehler_typ": None,
        "anzahl_fragen": 0,
        "erfolg": False,
    }

    # Kombinierte Analyse und Konvertierung (schneller, nur 1x PDF-Zugriff)
    try:
        meta, base64_bild = analysiere_und_konvertiere_seite(pdf_doc, seiten_nr)
    except Exception as e:
        fehler_typ, fehler_msg = erkenne_fehlertyp(None, e)
        metriken["fehler_typ"] = fehler_typ
        metriken["gesamtzeit_s"] = round(time.time() - seite_startzeit, 2)
        return None, fehler_msg, fehler_typ, 0, metriken

    # Strategie-Reihenfolge dynamisch
    if meta["ist_bildlastig"]:
        strategien = [
            ("bildlastig", get_bildlastig_prompt),
            ("megaprompt", get_megaprompt_mit_quelle),
            ("standard", get_standard_prompt),
        ]
    else:
        strategien = [
            ("megaprompt", get_megaprompt_mit_quelle),
            ("standard", get_standard_prompt),
            ("bildlastig", get_bildlastig_prompt),
        ]

    for strategie_idx, (strategie_name, prompt_func) in enumerate(strategien):
        for versuch in range(MAX_VERSUCHE):
            try:
                if versuch == 0:
                    print(f"  [{strategie_name.upper()}] Versuch {versuch + 1}...")
                else:
                    print(f"  [{strategie_name.upper()}] Wiederholung {versuch + 1}/{MAX_VERSUCHE}...")

                prompt = prompt_func(seiten_info)
                call_start = time.time()
                antwort, usage = rufe_openai_vision(client, base64_bild, prompt, modell)
                call_dauer = round(time.time() - call_start, 2)
                api_aufrufe += 1
                letzte_antwort = antwort

                csv_inhalt, anzahl_fragen, validierungs_fehler = bereinige_und_validiere_csv(antwort)

                if validierungs_fehler:
                    print(f"  ⚠️  {len(validierungs_fehler)} Validierungsprobleme")

                if anzahl_fragen == 0:
                    # Strategie wechseln
                    if strategie_idx < len(strategien) - 1:
                        print("  → Keine gültigen Fragen, wechsle Strategie...")
                        break
                    metriken["api_aufrufe"] = api_aufrufe
                    metriken["gesamtzeit_s"] = round(time.time() - seite_startzeit, 2)
                    metriken["fehler_typ"] = "UNGUELTIG_CSV"
                    return None, "Keine gültigen CSV-Zeilen generiert", "UNGUELTIG_CSV", 0, metriken

                # Erfolg!
                metriken["strategie"] = strategie_name
                metriken["api_aufrufe"] = api_aufrufe
                metriken["tokens_eingabe"] = usage.prompt_tokens if usage else 0
                metriken["tokens_ausgabe"] = usage.completion_tokens if usage else 0
                metriken["tokens_gesamt"] = usage.total_tokens if usage else 0
                metriken["antwortzeit_s"] = call_dauer
                metriken["gesamtzeit_s"] = round(time.time() - seite_startzeit, 2)
                metriken["anzahl_fragen"] = anzahl_fragen
                metriken["erfolg"] = True

                print(f"  ✓ {anzahl_fragen} gültige Fragen")
                return csv_inhalt, None, None, anzahl_fragen, metriken

            except Exception as e:
                api_aufrufe += 1
                fehler_typ, fehler_msg = erkenne_fehlertyp(letzte_antwort, e)

                # Retry bei technischen Fehlern
                if fehler_typ in ["VERBINDUNG", "RATE_LIMIT", "ZEITÜBERSCHREITUNG", "API_FEHLER"]:
                    if versuch < MAX_VERSUCHE - 1:
                        wartezeit = berechne_wartezeit(versuch, fehler_typ)
                        print(f"  ⚠️  {fehler_typ}: {str(e)[:80]}...")
                        print(f"  ⏳ Warte {wartezeit:.1f}s vor nächstem Versuch...")
                        time.sleep(wartezeit)
                        continue
                    else:
                        print(f"  ❌ {fehler_typ}: Max. Versuche erreicht für {strategie_name}")
                        metriken["api_aufrufe"] = api_aufrufe
                        metriken["gesamtzeit_s"] = round(time.time() - seite_startzeit, 2)
                        metriken["fehler_typ"] = fehler_typ
                        return None, fehler_msg, fehler_typ, 0, metriken

                # Nicht-technischer Fehler: Strategie wechseln
                print(f"  ❌ {fehler_typ}: {str(e)[:80]}")
                if strategie_idx < len(strategien) - 1:
                    print("  → Wechsle zu nächster Strategie...")
                    time.sleep(2)
                    break

    metriken["api_aufrufe"] = api_aufrufe
    metriken["gesamtzeit_s"] = round(time.time() - seite_startzeit, 2)
    metriken["fehler_typ"] = "UNBEKANNT"
    return None, "Alle Strategien fehlgeschlagen", "UNBEKANNT", 0, metriken


def verarbeite_seite_wrapper(args: tuple) -> dict:
    """Wrapper für parallele Verarbeitung."""
    client, pdf_pfad, seiten_nr, seiten_info, modell, fehler_log_pfad, pdf_seite, gesamt_seiten = args

    ergebnis = {
        "pdf_seite": pdf_seite,
        "seiten_info": seiten_info,
        "csv_inhalt": None,
        "fehler": None,
        "fehler_typ": None,
        "anzahl_fragen": 0,
        "metriken": {}
    }

    try:
        # PDF für diesen Thread öffnen
        pdf_doc = fitz.open(pdf_pfad)
        csv_inhalt, fehler, fehler_typ, anzahl, metriken = verarbeite_seite(
            client, pdf_doc, seiten_nr, seiten_info, modell, fehler_log_pfad
        )
        pdf_doc.close()
        ergebnis["csv_inhalt"] = csv_inhalt
        ergebnis["fehler"] = fehler
        ergebnis["fehler_typ"] = fehler_typ
        ergebnis["anzahl_fragen"] = anzahl
        ergebnis["metriken"] = metriken
    except Exception as e:
        ergebnis["fehler"] = str(e)
        ergebnis["fehler_typ"] = "UNBEKANNT"
        ergebnis["metriken"] = {"api_aufrufe": 0, "erfolg": False, "fehler_typ": "UNBEKANNT"}

    return ergebnis


def schreibe_protokoll_zeile(
    protokoll_pfad: str,
    pdf_pfad: str,
    prompt_pfad: str,
    modell: str,
    start_seite: int,
    end_seite: int,
    statistik: dict,
    metriken_liste: list,
    dauer_s: float,
    parallel: int,
    single_page: bool,
    unterbrochen: bool,
    ausgabe_pfad: str,
    metriken_pfad: str,
):
    """
    Fügt eine Zeile in das zentrale Excel-Protokoll ein.
    Erstellt die Datei mit formatierter Kopfzeile wenn sie noch nicht existiert.
    Erfordert: pip install openpyxl
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    SPALTEN = [
        "Datum", "Uhrzeit", "PDF", "Megaprompt", "Modell",
        "Seiten_Bereich", "Seiten_Gesamt", "Seiten_Erfolg", "Seiten_Fehler", "Erfolgsrate_%",
        "Fragen_Gesamt", "Fragen_pro_Seite", "API_Aufrufe_Gesamt",
        "Tokens_Eingabe", "Tokens_Ausgabe", "Tokens_Gesamt",
        "Dauer_Minuten", "Zeit_pro_Seite_s", "Parallelität", "Seitenmodus",
        "Dominante_Strategie", "Fehler_nach_Typ", "Unterbrochen",
        "CSV_Pfad", "Metriken_Pfad",
    ]

    SPALTENBREITEN = {
        "Datum": 14, "Uhrzeit": 12, "PDF": 35, "Megaprompt": 35, "Modell": 12,
        "Seiten_Bereich": 14, "Seiten_Gesamt": 14, "Seiten_Erfolg": 14,
        "Seiten_Fehler": 14, "Erfolgsrate_%": 13, "Fragen_Gesamt": 14,
        "Fragen_pro_Seite": 16, "API_Aufrufe_Gesamt": 18,
        "Tokens_Eingabe": 16, "Tokens_Ausgabe": 16, "Tokens_Gesamt": 14,
        "Dauer_Minuten": 15, "Zeit_pro_Seite_s": 16, "Parallelität": 13,
        "Seitenmodus": 13, "Dominante_Strategie": 20, "Fehler_nach_Typ": 40,
        "Unterbrochen": 13, "CSV_Pfad": 50, "Metriken_Pfad": 50,
    }

    # Werte berechnen
    verarbeitet = statistik["erfolg"] + statistik["fehlgeschlagen"]
    erfolgsrate = round(100 * statistik["erfolg"] / max(1, verarbeitet), 1)
    fragen_pro_seite = round(statistik["fragen_generiert"] / max(1, statistik["erfolg"]), 1)
    tokens_ein = sum(m.get("tokens_eingabe", 0) for m in metriken_liste)
    tokens_aus = sum(m.get("tokens_ausgabe", 0) for m in metriken_liste)
    tokens_ges = sum(m.get("tokens_gesamt", 0) for m in metriken_liste)
    api_aufrufe = sum(m.get("api_aufrufe", 0) for m in metriken_liste)
    zeit_pro_seite = round(dauer_s / max(1, verarbeitet), 1)

    # Dominante Strategie
    strategie_zaehler = {}
    for m in metriken_liste:
        s = m.get("strategie")
        if s:
            strategie_zaehler[s] = strategie_zaehler.get(s, 0) + 1
    dominante_strategie = max(strategie_zaehler, key=strategie_zaehler.get) if strategie_zaehler else "—"

    # Fehler nach Typ
    fehler_str = " | ".join(
        f"{typ}: {anz}" for typ, anz in statistik.get("fehler_nach_typ", {}).items()
    ) or "—"

    jetzt = datetime.now()
    zeilen_daten = [
        jetzt.strftime("%Y-%m-%d"),
        jetzt.strftime("%H:%M:%S"),
        Path(pdf_pfad).name,
        Path(prompt_pfad).name,
        modell,
        f"{start_seite}–{end_seite}",
        statistik["gesamt_seiten"],
        statistik["erfolg"],
        statistik["fehlgeschlagen"],
        erfolgsrate,
        statistik["fragen_generiert"],
        fragen_pro_seite,
        api_aufrufe,
        tokens_ein,
        tokens_aus,
        tokens_ges,
        round(dauer_s / 60, 1),
        zeit_pro_seite,
        parallel,
        "Einzel" if single_page else "Doppel",
        dominante_strategie,
        fehler_str,
        "JA" if unterbrochen else "NEIN",
        ausgabe_pfad,
        metriken_pfad,
    ]

    # Excel öffnen oder neu erstellen
    if os.path.exists(protokoll_pfad):
        wb = load_workbook(protokoll_pfad)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Protokoll"

        # Kopfzeile mit blauem Hintergrund
        for col_idx, spalte in enumerate(SPALTEN, 1):
            zelle = ws.cell(row=1, column=col_idx, value=spalte)
            zelle.font = Font(bold=True, color="FFFFFF")
            zelle.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            zelle.alignment = Alignment(horizontal="center", wrap_text=False)

        # Spaltenbreiten setzen
        for col_idx, spalte in enumerate(SPALTEN, 1):
            breite = SPALTENBREITEN.get(spalte, 15)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = breite

        # Erste Zeile einfrieren (Kopfzeile bleibt beim Scrollen sichtbar)
        ws.freeze_panes = "A2"

    # Datenzeile einfügen
    zeile_nr = ws.max_row + 1
    for col_idx, wert in enumerate(zeilen_daten, 1):
        ws.cell(row=zeile_nr, column=col_idx, value=wert)

    wb.save(protokoll_pfad)
    print(f"   Protokoll: {protokoll_pfad} (Zeile {zeile_nr - 1})")


def main():
    global MEGAPROMPT_INHALT, PROGRESS_PFAD, ERLEDIGTE_SEITEN, METRIKEN_LISTE

    parser = argparse.ArgumentParser(
        description="Konvertiert ein PDF in Fragen-CSV via OpenAI Vision",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument("--pdf", type=str, help="Pfad zur PDF-Datei")
    parser.add_argument("--prompt", type=str, help="Pfad zur Megaprompt DOCX-Datei")
    parser.add_argument("--start", type=int, default=1, help="Startseite (1-indiziert, Standard: 1)")
    parser.add_argument("--end", type=int, default=None, help="Endseite (inklusive, Standard: letzte Seite)")
    parser.add_argument("--output", type=str, default=None, help="Ausgabe-CSV-Datei")
    parser.add_argument("--test", action="store_true", help="Testmodus: nur 2 Seiten")
    parser.add_argument("--book-start", type=int, default=None, help="Erste Buchseitennummer (z.B. wenn PDF-Seite 8 = Buchseite 15, dann --book-start 15)")
    parser.add_argument("--single-page", action="store_true", help="1 PDF-Seite = 1 Buchseite (Standard: 1 PDF-Seite = 2 Buchseiten/Doppelseite)")
    parser.add_argument(
        "--model",
        type=str,
        default="gpt-4.1",
        choices=["gpt-4.1", "gpt-5.1", "gpt-5.2"],
        help="OpenAI-Modell (Standard: gpt-4.1)"
    )
    parser.add_argument("--parallel", type=int, default=3, help="Anzahl paralleler Anfragen (Standard: 3)")

    args = parser.parse_args()

    # Megaprompt-Pfad
    prompt_pfad = args.prompt
    if not prompt_pfad:
        prompt_pfad = input("📝 Pfad zur Megaprompt DOCX-Datei: ").strip().strip('"')

    # Megaprompt laden
    print(f"\n📝 Lade Megaprompt: {prompt_pfad}")
    try:
        MEGAPROMPT_INHALT = lade_megaprompt(prompt_pfad)
        print(f"   ✓ Megaprompt geladen ({len(MEGAPROMPT_INHALT)} Zeichen)")
    except FileNotFoundError as e:
        print(f"❌ FEHLER: {e}")
        sys.exit(1)

    # API-Schlüssel prüfen
    api_schluessel = os.getenv("OPENAI_API_KEY")
    if not api_schluessel:
        print("❌ FEHLER: OPENAI_API_KEY nicht gefunden!")
        print("   Erstellen Sie eine .env-Datei mit:")
        print("   OPENAI_API_KEY=sk-ihr-api-schluessel")
        sys.exit(1)

    # PDF-Pfad
    pdf_pfad = args.pdf
    if not pdf_pfad:
        pdf_pfad = input("📁 Pfad zur PDF-Datei: ").strip().strip('"')

    if not os.path.exists(pdf_pfad):
        print(f"❌ FEHLER: Datei '{pdf_pfad}' nicht gefunden!")
        sys.exit(1)

    # Seitenanzahl
    gesamt_seiten = hole_pdf_seitenanzahl(pdf_pfad)
    print(f"\n📄 PDF: {pdf_pfad}")
    print(f"   Seiten: {gesamt_seiten}")

    # Bereich
    start_seite = max(1, args.start)
    end_seite = args.end if args.end else gesamt_seiten
    end_seite = min(end_seite, gesamt_seiten)

    if start_seite > end_seite:
        print(f"❌ FEHLER: Startseite ({start_seite}) > Endseite ({end_seite})")
        sys.exit(1)

    if args.test:
        end_seite = min(start_seite + 1, end_seite)
        print("🧪 TESTMODUS: max. 2 Seiten")

    anzahl_seiten = end_seite - start_seite + 1
    print(f"   Zu verarbeiten: PDF-Seiten {start_seite}-{end_seite} ({anzahl_seiten} Seiten)")
    print(f"   Modell: {args.model}")
    print(f"   Parallel: {args.parallel} Thread(s)")
    print(f"   Seitenmodus: {'Einzelseiten (1 PDF = 1 Buchseite)' if args.single_page else 'Doppelseiten (1 PDF = 2 Buchseiten)'}")

    # Ausgabedatei
    if args.output:
        ausgabe_pfad = args.output
    else:
        pdf_name = Path(pdf_pfad).stem
        ausgabe_pfad = f"{pdf_name}_questions.csv"

    fehler_log_pfad = ausgabe_pfad.replace(".csv", "_errors.json")
    PROGRESS_PFAD = ausgabe_pfad.replace(".csv", "_progress.json")
    metriken_pfad = ausgabe_pfad.replace(".csv", "_metriken.json")
    print(f"\n📁 Ausgabe: {ausgabe_pfad}")
    print(f"   Fehler-Log: {fehler_log_pfad}")

    # Resume: bereits erledigte Seiten laden
    ERLEDIGTE_SEITEN = lade_progress(PROGRESS_PFAD)
    if ERLEDIGTE_SEITEN:
        print(f"\n🔄 RESUME ERKANNT: {len(ERLEDIGTE_SEITEN)} Seiten bereits verarbeitet → werden übersprungen")

    # Buchseitennummer - Standard: PDF-Seite = Buchseite
    if args.book_start:
        buch_start = args.book_start
    else:
        buch_start = start_seite  # PDF-Seite 8 = Buchseite 8 (einfacher Standard)

    # Modus für Seitenzählung
    seiten_pro_pdf = 1 if args.single_page else 2  # 1 = einzelne Seiten, 2 = Doppelseiten
    print(f"   Erste Buchseite: {buch_start} (PDF-Seite {start_seite} = Buchseite {buch_start})")

    # OpenAI-Client
    client = OpenAI(api_key=api_schluessel)

    # Verbindungstest
    print("\n🔌 Teste Verbindung zur OpenAI API...")
    try:
        test_response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role": "user", "content": "Sage nur: OK"}],
            max_tokens=5,
            timeout=30
        )
        _ = test_response.choices[0].message.content
        print("   ✓ Verbindung erfolgreich!")
    except Exception as e:
        print(f"   ❌ Verbindungsfehler: {type(e).__name__}")
        print(f"      {str(e)[:200]}")
        print("\n   Mögliche Ursachen:")
        print("   - Firewall blockiert api.openai.com")
        print("   - Proxy-Einstellungen erforderlich")
        print("   - VPN-Verbindung aktiv")
        print("   - API-Schlüssel ungültig")
        print("\n   Prüfen Sie: HTTPS_PROXY / HTTP_PROXY Umgebungsvariablen")
        antwort = input("\n   Trotzdem fortfahren? (j/n): ").strip().lower()
        if antwort != "j":
            sys.exit(1)

    # Statistiken
    statistik = {
        "gesamt_seiten": anzahl_seiten,
        "erfolg": 0,
        "fehlgeschlagen": 0,
        "fragen_generiert": 0,
        "fehler_nach_typ": {}
    }

    # Aufgaben vorbereiten (bereits erledigte Seiten überspringen)
    aufgaben = []
    for seiten_nr in range(start_seite - 1, end_seite):
        pdf_seite = seiten_nr + 1

        if pdf_seite in ERLEDIGTE_SEITEN:
            continue  # Resume: Seite bereits verarbeitet

        seiten_index = seiten_nr - (start_seite - 1)

        if seiten_pro_pdf == 1:
            # Einzelseiten-Modus: 1 PDF-Seite = 1 Buchseite
            buch_seite = buch_start + seiten_index
            seiten_info = f"Buch S. {buch_seite}"
        else:
            # Doppelseiten-Modus: 1 PDF-Seite = 2 Buchseiten
            buch_seite_links = buch_start + seiten_index * 2
            buch_seite_rechts = buch_seite_links + 1
            seiten_info = f"Buch S. {buch_seite_links}-{buch_seite_rechts}"

        aufgaben.append((
            client, pdf_pfad, seiten_nr, seiten_info,
            args.model, fehler_log_pfad, pdf_seite, gesamt_seiten
        ))

    # Kopfzeile schreiben
    datei_existiert = os.path.exists(ausgabe_pfad) and os.path.getsize(ausgabe_pfad) > 0
    with open(ausgabe_pfad, "a", encoding="utf-8-sig", newline="") as csv_datei:
        if not datei_existiert:
            csv_datei.write(CSV_HEADER + "\n")
            print("\n✓ CSV-Kopfzeile geschrieben")

    print(f"\n{'=' * 60}")
    print("VERARBEITUNG GESTARTET")
    print("(Ctrl+C zum Abbrechen - Daten werden gespeichert)")
    print(f"{'=' * 60}")

    # Signal-Handler für Ctrl+C registrieren
    signal.signal(signal.SIGINT, signal_handler)

    startzeit = time.time()
    fehler_liste = []

    if args.parallel > 1:
        # Parallele Verarbeitung: Ergebnisse SOFORT speichern (nicht am Ende)
        try:
            with ThreadPoolExecutor(max_workers=args.parallel) as executor:
                futures = {executor.submit(verarbeite_seite_wrapper, aufgabe): aufgabe[6] for aufgabe in aufgaben}

                for future in as_completed(futures):
                    ergebnis = future.result()
                    pdf_seite = ergebnis["pdf_seite"]
                    seiten_info = ergebnis["seiten_info"]
                    print(f"\n[Seite {pdf_seite}] {seiten_info}")

                    # Metriken sammeln
                    if ergebnis.get("metriken"):
                        m = ergebnis["metriken"]
                        m["pdf_seite"] = pdf_seite
                        m["seiten_info"] = seiten_info
                        METRIKEN_LISTE.append(m)

                    if ergebnis["fehler"]:
                        statistik["fehlgeschlagen"] += 1
                        fehler_typ = ergebnis["fehler_typ"] or "UNBEKANNT"
                        statistik["fehler_nach_typ"][fehler_typ] = statistik["fehler_nach_typ"].get(fehler_typ, 0) + 1
                        fehler_liste.append(f"Seite {pdf_seite}: {ergebnis['fehler']}")
                        protokolliere_fehler(fehler_log_pfad, pdf_seite, seiten_info, fehler_typ, ergebnis["fehler"])
                        print(f"  ❌ {fehler_typ}: {ergebnis['fehler'][:80]}...")
                    else:
                        # SOFORT in CSV schreiben (nicht am Ende)
                        schreibe_ergebnis_sofort(ausgabe_pfad, ergebnis, statistik)
                        print(f"  ✅ {ergebnis['anzahl_fragen']} Fragen generiert (gespeichert)")

                    # Abbruch NACH Verarbeitung prüfen — so wird die letzte Seite noch gespeichert
                    if stop_event.is_set():
                        for f in futures:
                            f.cancel()
                        break

        except Exception as e:
            print(f"\n❌ Fehler bei paralleler Verarbeitung: {e}")

    else:
        # Sequentielle Verarbeitung - PDF einmal öffnen
        pdf_doc = fitz.open(pdf_pfad)
        try:
            for aufgabe in aufgaben:
                if stop_event.is_set():
                    break

                pdf_seite = aufgabe[6]
                seiten_info = aufgabe[3]

                print(f"\n{'=' * 60}")
                print(f"Seite {pdf_seite}/{gesamt_seiten} ({seiten_info})")
                print(f"{'=' * 60}")

                csv_inhalt, fehler, fehler_typ, anzahl, metriken = verarbeite_seite(
                    client, pdf_doc, aufgabe[2], seiten_info, args.model, fehler_log_pfad
                )

                # Metriken sammeln
                metriken["pdf_seite"] = pdf_seite
                metriken["seiten_info"] = seiten_info
                METRIKEN_LISTE.append(metriken)

                if fehler:
                    statistik["fehlgeschlagen"] += 1
                    fehler_typ = fehler_typ or "UNBEKANNT"
                    statistik["fehler_nach_typ"][fehler_typ] = statistik["fehler_nach_typ"].get(fehler_typ, 0) + 1
                    fehler_liste.append(f"Seite {pdf_seite} ({seiten_info}): [{fehler_typ}] {fehler}")
                    protokolliere_fehler(fehler_log_pfad, pdf_seite, seiten_info, fehler_typ, fehler)
                    print(f"\n❌ FEHLGESCHLAGEN: {fehler_typ}")
                    print(f"   {fehler}")
                else:
                    statistik["erfolg"] += 1
                    statistik["fragen_generiert"] += anzahl
                    with open(ausgabe_pfad, "a", encoding="utf-8-sig", newline="") as csv_datei:
                        csv_datei.write(csv_inhalt + "\n")
                    if PROGRESS_PFAD:
                        speichere_progress(PROGRESS_PFAD, pdf_seite)
                    print(f"\n✅ OK: {anzahl} Fragen generiert (gespeichert)")
        finally:
            pdf_doc.close()

    # Zusammenfassung
    dauer = time.time() - startzeit

    print(f"\n{'=' * 60}")
    if wurde_unterbrochen:
        print("VERARBEITUNG UNTERBROCHEN - DATEN WURDEN GESPEICHERT")
    else:
        print("VERARBEITUNG ABGESCHLOSSEN")
    print(f"{'=' * 60}")

    verarbeitete_seiten = statistik["erfolg"] + statistik["fehlgeschlagen"]
    print("\n📊 STATISTIK:")
    print(f"   Dauer: {dauer/60:.1f} Minuten ({dauer:.0f} Sekunden)")
    if wurde_unterbrochen:
        print(f"   Seiten verarbeitet: {verarbeitete_seiten} von {statistik['gesamt_seiten']} (unterbrochen)")
    else:
        print(f"   Seiten verarbeitet: {statistik['gesamt_seiten']}")
    print(f"   Erfolgreich: {statistik['erfolg']} ({100*statistik['erfolg']/max(1,verarbeitete_seiten):.1f}%)")
    print(f"   Fehlgeschlagen: {statistik['fehlgeschlagen']} ({100*statistik['fehlgeschlagen']/max(1,verarbeitete_seiten):.1f}%)")
    print(f"   Fragen generiert: {statistik['fragen_generiert']}")
    if statistik["erfolg"] > 0:
        print(f"   Ø Fragen/Seite: {statistik['fragen_generiert']/statistik['erfolg']:.1f}")

    # Metriken speichern
    if METRIKEN_LISTE:
        gesamt_tokens = sum(m.get("tokens_gesamt", 0) for m in METRIKEN_LISTE)
        gesamt_api_aufrufe = sum(m.get("api_aufrufe", 0) for m in METRIKEN_LISTE)
        metriken_export = {
            "lauf_info": {
                "pdf": pdf_pfad,
                "megaprompt": prompt_pfad,
                "modell": args.model,
                "start_seite": start_seite,
                "end_seite": end_seite,
                "parallel": args.parallel,
                "startzeit": datetime.fromtimestamp(startzeit).isoformat(),
                "endzeit": datetime.now().isoformat(),
                "dauer_s": round(dauer, 2),
                "unterbrochen": wurde_unterbrochen,
            },
            "zusammenfassung": {
                "seiten_gesamt": statistik["gesamt_seiten"],
                "seiten_erfolg": statistik["erfolg"],
                "seiten_fehler": statistik["fehlgeschlagen"],
                "fragen_gesamt": statistik["fragen_generiert"],
                "api_aufrufe_gesamt": gesamt_api_aufrufe,
                "tokens_gesamt": gesamt_tokens,
            },
            "seiten": sorted(METRIKEN_LISTE, key=lambda x: x.get("pdf_seite", 0)),
        }
        with open(metriken_pfad, "w", encoding="utf-8") as f:
            json.dump(metriken_export, f, ensure_ascii=False, indent=2)
        print(f"\n📊 Metriken gespeichert: {metriken_pfad}")
        print(f"   API-Aufrufe gesamt: {gesamt_api_aufrufe}")
        print(f"   Tokens gesamt: {gesamt_tokens:,}")

    print("\n📁 DATEIEN:")
    print(f"   CSV: {ausgabe_pfad}")

    if fehler_liste:
        print(f"   Fehler-Log: {fehler_log_pfad}")
        print("\n⚠️  FEHLER NACH TYP:")
        for f_typ, anzahl in statistik["fehler_nach_typ"].items():
            print(f"   - {f_typ}: {anzahl} Seite(n)")
    elif not wurde_unterbrochen:
        print("\n✅ Keine Fehler!")

    # Protokoll-Zeile ins Excel schreiben
    protokoll_pfad = os.path.join(os.path.dirname(os.path.abspath(__file__)), "protokoll.xlsx")
    try:
        schreibe_protokoll_zeile(
            protokoll_pfad=protokoll_pfad,
            pdf_pfad=pdf_pfad,
            prompt_pfad=prompt_pfad,
            modell=args.model,
            start_seite=start_seite,
            end_seite=end_seite,
            statistik=statistik,
            metriken_liste=METRIKEN_LISTE,
            dauer_s=dauer,
            parallel=args.parallel,
            single_page=args.single_page,
            unterbrochen=wurde_unterbrochen,
            ausgabe_pfad=ausgabe_pfad,
            metriken_pfad=metriken_pfad,
        )
    except ImportError:
        print("\n   ⚠️  Protokoll nicht geschrieben — bitte 'pip install openpyxl' ausführen")
    except PermissionError:
        print("\n   ⚠️  Protokoll gesperrt — bitte protokoll.xlsx in Excel schließen und nochmals starten")
    except Exception as e:
        print(f"\n   ⚠️  Protokoll-Fehler: {e}")

    if wurde_unterbrochen:
        print(f"\n💡 Zum Fortsetzen: Starte das Programm mit denselben Parametern neu.")
        print(f"   Das Resume-System springt automatisch zu den noch nicht verarbeiteten Seiten.")
    else:
        loesche_progress(PROGRESS_PFAD)


if __name__ == "__main__":
    main()
